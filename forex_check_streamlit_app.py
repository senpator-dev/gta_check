#!/usr/bin/env python3
"""Сводит FB-выгрузку расходов и MVP-выгрузку количеств в текстовый чек.

Базовый запуск из папки с выгрузками:
    python forex_check.py

Явно указать файлы:
    python forex_check.py --fb "Статистика аккаунтов (80).xlsx" --mvp "data (85).xlsx"

Сохранить чек в файл и одновременно вывести в консоль:
    python forex_check.py --output check.txt

Посчитать общую стату по фильтру из FB-выгрузки:
    python forex_check.py --stats "SPAIN"
    python forex_check.py --stats "Общая стата по SPAIN"

Работать по объявлениям вместо кампаний:
    python forex_check.py --entity ad

По умолчанию:
- в чек попадает каждая FB-строка выбранного уровня с ненулевым спендом;
- если FB-строки нет в MVP, её цены ПДП и диалога равны 0;
- лишние пробелы в названиях игнорируются только при матчинге;
- бюджеты добавляются рядом с названием в квадратных скобках, например [100$];
- после каждой страны выводится общий чек по стране: Общ.: $xx.xx / $xx.xx / $xx.xx;
- --stats считает клики, просмотры и расход по точному блоку названия, разделённому через " - ":
  фильтр SPAIN совпадёт с "... - SPAIN - ...", но НЕ совпадёт с "... - SPAIN-ES - ...";
- --entity auto сам выбирает объявления, если в FB-таблице есть колонка объявления, иначе кампании.
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Iterable, Iterator, Sequence

from openpyxl import load_workbook

MONEY_QUANT = Decimal("0.01")
ZERO = Decimal("0")

# Приоритет из эталона. Гео, которых нет в этом списке, идут после него
# по порядку первого появления в FB-выгрузке.
GEO_PRIORITY = ("Бельгия", "Чехия", "Испания")

# Полные русские названия для ISO alpha-2. Можно расширять алиасы ниже,
# если в названиях кампаний/объявлений используется не ISO-код, а внутреннее сокращение.
ISO2_TO_RU = {
    "AD": "Андорра", "AE": "ОАЭ", "AF": "Афганистан", "AG": "Антигуа и Барбуда",
    "AL": "Албания", "AM": "Армения", "AO": "Ангола", "AR": "Аргентина",
    "AT": "Австрия", "AU": "Австралия", "AZ": "Азербайджан", "BA": "Босния и Герцеговина",
    "BB": "Барбадос", "BD": "Бангладеш", "BE": "Бельгия", "BF": "Буркина-Фасо",
    "BG": "Болгария", "BH": "Бахрейн", "BI": "Бурунди", "BJ": "Бенин",
    "BN": "Бруней", "BO": "Боливия", "BR": "Бразилия", "BS": "Багамы",
    "BT": "Бутан", "BW": "Ботсвана", "BY": "Беларусь", "BZ": "Белиз",
    "CA": "Канада", "CD": "ДР Конго", "CF": "ЦАР", "CG": "Республика Конго",
    "CH": "Швейцария", "CI": "Кот-д’Ивуар", "CL": "Чили", "CM": "Камерун",
    "CN": "Китай", "CO": "Колумбия", "CR": "Коста-Рика", "CU": "Куба",
    "CV": "Кабо-Верде", "CY": "Кипр", "CZ": "Чехия", "DE": "Германия",
    "DJ": "Джибути", "DK": "Дания", "DM": "Доминика", "DO": "Доминиканская Республика",
    "DZ": "Алжир", "EC": "Эквадор", "EE": "Эстония", "EG": "Египет",
    "ER": "Эритрея", "ES": "Испания", "ET": "Эфиопия", "FI": "Финляндия",
    "FJ": "Фиджи", "FM": "Микронезия", "FR": "Франция", "GA": "Габон",
    "GB": "Великобритания", "GD": "Гренада", "GE": "Грузия", "GH": "Гана",
    "GM": "Гамбия", "GN": "Гвинея", "GQ": "Экваториальная Гвинея", "GR": "Греция",
    "GT": "Гватемала", "GW": "Гвинея-Бисау", "GY": "Гайана", "HN": "Гондурас",
    "HR": "Хорватия", "HT": "Гаити", "HU": "Венгрия", "ID": "Индонезия",
    "IE": "Ирландия", "IL": "Израиль", "IN": "Индия", "IQ": "Ирак",
    "IR": "Иран", "IS": "Исландия", "IT": "Италия", "JM": "Ямайка",
    "JO": "Иордания", "JP": "Япония", "KE": "Кения", "KG": "Кыргызстан",
    "KH": "Камбоджа", "KI": "Кирибати", "KM": "Коморы", "KN": "Сент-Китс и Невис",
    "KP": "КНДР", "KR": "Южная Корея", "KW": "Кувейт", "KZ": "Казахстан",
    "LA": "Лаос", "LB": "Ливан", "LC": "Сент-Люсия", "LI": "Лихтенштейн",
    "LK": "Шри-Ланка", "LR": "Либерия", "LS": "Лесото", "LT": "Литва",
    "LU": "Люксембург", "LV": "Латвия", "LY": "Ливия", "MA": "Марокко",
    "MC": "Монако", "MD": "Молдова", "ME": "Черногория", "MG": "Мадагаскар",
    "MH": "Маршалловы Острова", "MK": "Северная Македония", "ML": "Мали", "MM": "Мьянма",
    "MN": "Монголия", "MR": "Мавритания", "MT": "Мальта", "MU": "Маврикий",
    "MV": "Мальдивы", "MW": "Малави", "MX": "Мексика", "MY": "Малайзия",
    "MZ": "Мозамбик", "NA": "Намибия", "NE": "Нигер", "NG": "Нигерия",
    "NI": "Никарагуа", "NL": "Нидерланды", "NO": "Норвегия", "NP": "Непал",
    "NR": "Науру", "NZ": "Новая Зеландия", "OM": "Оман", "PA": "Панама",
    "PE": "Перу", "PG": "Папуа — Новая Гвинея", "PH": "Филиппины", "PK": "Пакистан",
    "PL": "Польша", "PT": "Португалия", "PW": "Палау", "PY": "Парагвай",
    "QA": "Катар", "RO": "Румыния", "RS": "Сербия", "RU": "Россия",
    "RW": "Руанда", "SA": "Саудовская Аравия", "SB": "Соломоновы Острова", "SC": "Сейшелы",
    "SD": "Судан", "SE": "Швеция", "SG": "Сингапур", "SI": "Словения",
    "SK": "Словакия", "SL": "Сьерра-Леоне", "SM": "Сан-Марино", "SN": "Сенегал",
    "SO": "Сомали", "SR": "Суринам", "SS": "Южный Судан", "ST": "Сан-Томе и Принсипи",
    "SV": "Сальвадор", "SY": "Сирия", "SZ": "Эсватини", "TD": "Чад",
    "TG": "Того", "TH": "Таиланд", "TJ": "Таджикистан", "TL": "Восточный Тимор",
    "TM": "Туркменистан", "TN": "Тунис", "TO": "Тонга", "TR": "Турция",
    "TT": "Тринидад и Тобаго", "TV": "Тувалу", "TZ": "Танзания", "UA": "Украина",
    "UG": "Уганда", "US": "США", "UY": "Уругвай", "UZ": "Узбекистан",
    "VA": "Ватикан", "VC": "Сент-Винсент и Гренадины", "VE": "Венесуэла", "VN": "Вьетнам",
    "VU": "Вануату", "WS": "Самоа", "YE": "Йемен", "ZA": "ЮАР", "ZM": "Замбия",
    "ZW": "Зимбабве",
}

# Внутренние обозначения, которые встречались в рабочих выгрузках.
GEO_ALIASES = {
    "SPAIN": "Испания", "GERM": "Германия", "GERMANY": "Германия",
    "POLAND": "Польша", "BELGIUM": "Бельгия", "CZECH": "Чехия",
    "CZECHIA": "Чехия", "KYRGYZSTAN": "Кыргызстан", "KYRGYZ": "Кыргызстан",
    **ISO2_TO_RU,
}

# Алиасы заголовков FB. При необходимости просто добавляй сюда новые варианты.
FB_CAMPAIGN_HEADERS = ("кампания", "название кампании", "campaign", "campaign name")
FB_AD_HEADERS = ("объявление", "название объявления", "ad", "ad name")
FB_SPEND_HEADERS = (
    "расход", "сумма затрат", "потраченная сумма", "amount spent", "amount spent (usd)", "spent", "spend"
)
FB_CLICK_HEADERS = (
    "клики", "клики (все)", "клики по ссылке", "link clicks", "clicks", "clicks (all)"
)
FB_VIEW_HEADERS = (
    "просмотры", "просмотры целевой страницы", "просмотры лендинга", "landing page views",
    "views", "показы", "impressions"
)

# Алиасы заголовков MVP. name оставлен в обоих уровнях, потому что MVP часто отдаёт универсальную колонку.
MVP_CAMPAIGN_HEADERS = ("name", "campaign", "campaign name", "кампания", "название кампании")
MVP_AD_HEADERS = ("name", "ad", "ad name", "объявление", "название объявления")
MVP_SUB_HEADERS = ("sub", "subs", "пдп", "подписки")
MVP_CHAT_HEADERS = ("chat", "chats", "диа", "диалоги")

ENTITY_LABELS = {
    "campaign": "кампаний",
    "ad": "объявлений",
}


@dataclass(frozen=True)
class FbItem:
    title: str
    normalized_title: str
    spend: Decimal
    budget: Decimal | None
    row_number: int
    first_seen_index: int
    clicks: Decimal | None = None
    views: Decimal | None = None


@dataclass(frozen=True)
class MvpItem:
    title: str
    normalized_title: str
    sub: Decimal
    chat: Decimal
    row_number: int


def normalize_spaces(value: object) -> str:
    """Схлопывает любые пробельные символы для матчинга, но не для вывода."""
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_header(value: object) -> str:
    return normalize_spaces(value).lower()


def parse_decimal(value: object, *, blank_is_zero: bool = False, field: str = "значение") -> Decimal:
    if value is None or normalize_spaces(value) == "":
        if blank_is_zero:
            return ZERO
        raise ValueError(f"Пустое поле: {field}")
    if isinstance(value, bool):
        raise ValueError(f"Некорректное числовое значение в поле {field}: {value!r}")
    if isinstance(value, (int, float, Decimal)):
        text = str(value)
    else:
        text = normalize_spaces(value)
        text = text.replace("$", "").replace("USD", "").replace("usd", "")
        text = text.replace(" ", "")
        # 1 234,56 -> 1234.56. Для одиночной запятой считаем её десятичным разделителем.
        if "," in text and "." not in text:
            text = text.replace(",", ".")
        elif "," in text and "." in text:
            text = text.replace(",", "")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"Не удалось прочитать число в поле {field}: {value!r}") from exc


def ensure_nonnegative_integer(value: Decimal, *, field: str, row_number: int) -> Decimal:
    if value < 0 or value != value.to_integral_value():
        raise ValueError(f"{field} должен быть целым неотрицательным числом, строка {row_number}: {value}")
    return value


def format_money(value: Decimal) -> str:
    rounded = value.quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)
    return f"${rounded:.2f}"


def format_metric(spend: Decimal, count: Decimal) -> str:
    return "0" if count == 0 else format_money(spend / count)


def format_plain_number(value: Decimal) -> str:
    if value == value.to_integral_value():
        return str(int(value))
    return str(value.quantize(MONEY_QUANT, rounding=ROUND_HALF_UP))


def normalize_aliases(aliases: Sequence[str]) -> tuple[str, ...]:
    return tuple(normalize_header(alias) for alias in aliases)


def iter_sheet_rows(path: Path) -> Iterator[tuple[str, list[tuple[object, ...]]]]:
    """Читает листы быстро; при нестандартном диапазоне Excel форсирует пересчёт dimensions."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            try:
                ws.calculate_dimension(force=True)
            except Exception:
                pass
            yield ws.title, list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        # Запасной режим для файлов с повреждённым/нестандартным dimension ref.
        wb = load_workbook(path, read_only=False, data_only=True)
        for ws in wb.worksheets:
            yield ws.title, list(ws.iter_rows(values_only=True))
        wb.close()


def row_header_positions(row: Sequence[object]) -> dict[str, int]:
    header_positions: dict[str, int] = {}
    for col_idx, value in enumerate(row):
        key = normalize_header(value)
        if key and key not in header_positions:
            header_positions[key] = col_idx
    return header_positions


def resolve_col(headers: dict[str, int], aliases: Sequence[str]) -> int | None:
    for alias in normalize_aliases(aliases):
        if alias in headers:
            return headers[alias]
    return None


def find_table_header(
    path: Path,
    *,
    name_candidates: Sequence[Sequence[str]],
    required_candidates: Sequence[Sequence[str]],
    scan_limit: int = 15,
) -> tuple[str, list[tuple[object, ...]], int, dict[str, int], int, int]:
    """Ищет строку заголовков и возвращает выбранную колонку имени.

    name_candidates идут по приоритету: например, при auto сначала объявления, потом кампании.
    Возвращает индекс выбранной группы name_candidates, чтобы понять resolved entity.
    """
    for sheet_name, rows in iter_sheet_rows(path):
        for idx, row in enumerate(rows[:scan_limit]):
            headers = row_header_positions(row)
            required_ok = all(resolve_col(headers, aliases) is not None for aliases in required_candidates)
            if not required_ok:
                continue
            for candidate_idx, aliases in enumerate(name_candidates):
                name_col = resolve_col(headers, aliases)
                if name_col is not None:
                    return sheet_name, rows, idx, headers, name_col, candidate_idx
    required = ", ".join("/".join(group) for group in required_candidates)
    names = ", ".join("/".join(group) for group in name_candidates)
    raise ValueError(f"В файле {path.name} не найдена строка заголовков. Нужны колонки имени ({names}) и колонки: {required}")


def cell(row: Sequence[object], index: int | None) -> object:
    if index is None or index >= len(row):
        return None
    return row[index]


def entity_name_candidates(entity: str, *, source: str) -> list[tuple[str, tuple[str, ...]]]:
    if source == "fb":
        campaign_headers = FB_CAMPAIGN_HEADERS
        ad_headers = FB_AD_HEADERS
    elif source == "mvp":
        campaign_headers = MVP_CAMPAIGN_HEADERS
        ad_headers = MVP_AD_HEADERS
    else:
        raise ValueError(f"Неизвестный источник: {source}")

    if entity == "campaign":
        return [("campaign", campaign_headers)]
    if entity == "ad":
        return [("ad", ad_headers)]
    if entity == "auto":
        # Если в FB-таблице есть колонка объявления — считаем чек по объявлениям.
        # Если нужна стата по кампаниям из ad-level выгрузки, запускай с --entity campaign.
        return [("ad", ad_headers), ("campaign", campaign_headers)]
    raise ValueError(f"Некорректный уровень: {entity}")


def strip_fb_name_tail(raw_value: object) -> tuple[str, Decimal | None]:
    raw = str(raw_value or "").strip()
    # Название — всё до первой скобки, перед которой есть пробел.
    title = re.split(r"\s+\(", raw, maxsplit=1)[0].rstrip()
    budget_match = re.search(
        r"Дневной\s+бюджет\s+(?:кампании|объявления)?\s*-\s*([\d\s.,]+)\s*USD",
        raw,
        flags=re.IGNORECASE,
    )
    budget = parse_decimal(budget_match.group(1), field="дневной бюджет") if budget_match else None
    return title, budget


def parse_fb(
    path: Path,
    *,
    entity: str = "auto",
    include_zero_spend: bool = False,
    require_stats: bool = False,
) -> tuple[list[FbItem], str]:
    name_options = entity_name_candidates(entity, source="fb")
    _, rows, header_idx, headers, name_col, chosen_idx = find_table_header(
        path,
        name_candidates=[headers for _, headers in name_options],
        required_candidates=(FB_SPEND_HEADERS,),
    )
    resolved_entity = name_options[chosen_idx][0]
    spend_col = resolve_col(headers, FB_SPEND_HEADERS)
    click_col = resolve_col(headers, FB_CLICK_HEADERS)
    view_col = resolve_col(headers, FB_VIEW_HEADERS)

    if spend_col is None:
        raise ValueError(f"В файле {path.name} не найдена колонка расхода")
    if require_stats and click_col is None:
        raise ValueError(
            f"В файле {path.name} не найдена колонка кликов. "
            f"Добавь её в выгрузку или расширь FB_CLICK_HEADERS в скрипте."
        )
    if require_stats and view_col is None:
        raise ValueError(
            f"В файле {path.name} не найдена колонка просмотров/показов. "
            f"Добавь её в выгрузку или расширь FB_VIEW_HEADERS в скрипте."
        )

    output: list[FbItem] = []
    seen_index = 0
    for excel_row_number, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        raw_name = cell(row, name_col)
        if raw_name is None or normalize_spaces(raw_name) == "":
            continue
        title, budget = strip_fb_name_tail(raw_name)
        if not title:
            continue
        spend = parse_decimal(cell(row, spend_col), blank_is_zero=True, field=f"Расход, строка {excel_row_number}")
        if spend == 0 and not include_zero_spend:
            continue
        clicks = None
        views = None
        if click_col is not None:
            clicks = parse_decimal(cell(row, click_col), blank_is_zero=True, field=f"Клики, строка {excel_row_number}")
        if view_col is not None:
            views = parse_decimal(cell(row, view_col), blank_is_zero=True, field=f"Просмотры, строка {excel_row_number}")
        output.append(
            FbItem(
                title=title,
                normalized_title=normalize_spaces(title),
                spend=spend,
                budget=budget,
                row_number=excel_row_number,
                first_seen_index=seen_index,
                clicks=clicks,
                views=views,
            )
        )
        seen_index += 1
    return output, resolved_entity


def parse_mvp(path: Path, *, entity: str) -> list[MvpItem]:
    name_options = entity_name_candidates(entity, source="mvp")
    _, rows, header_idx, headers, name_col, _ = find_table_header(
        path,
        name_candidates=[headers for _, headers in name_options],
        required_candidates=(MVP_SUB_HEADERS, MVP_CHAT_HEADERS),
    )
    sub_col = resolve_col(headers, MVP_SUB_HEADERS)
    chat_col = resolve_col(headers, MVP_CHAT_HEADERS)
    if sub_col is None or chat_col is None:
        raise ValueError(f"В файле {path.name} не найдены колонки sub/chat")

    output: list[MvpItem] = []
    for excel_row_number, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        raw_title = cell(row, name_col)
        if raw_title is None or normalize_spaces(raw_title) == "":
            continue
        title = str(raw_title).strip()
        sub = parse_decimal(cell(row, sub_col), blank_is_zero=True, field=f"sub, строка {excel_row_number}")
        chat = parse_decimal(cell(row, chat_col), blank_is_zero=True, field=f"chat, строка {excel_row_number}")
        output.append(
            MvpItem(
                title=title,
                normalized_title=normalize_spaces(title),
                sub=ensure_nonnegative_integer(sub, field="sub", row_number=excel_row_number),
                chat=ensure_nonnegative_integer(chat, field="chat", row_number=excel_row_number),
                row_number=excel_row_number,
            )
        )
    return output


def duplicate_titles(items: Iterable[FbItem | MvpItem]) -> list[str]:
    counts = Counter(item.normalized_title for item in items)
    return sorted(title for title, count in counts.items() if count > 1)


def detect_geo(title: str) -> str | None:
    normalized = normalize_spaces(title).upper()
    # Сначала алиасы длиной больше двух символов, затем ISO-коды.
    words = re.findall(r"[A-Z]{2,}", normalized)
    for token in words:
        if len(token) > 2 and token in GEO_ALIASES:
            return GEO_ALIASES[token]
    for token in words:
        if token in GEO_ALIASES:
            return GEO_ALIASES[token]
    return None


def auto_pick(directory: Path, patterns: Sequence[str]) -> Path:
    candidates: list[Path] = []
    for pattern in patterns:
        candidates.extend(directory.glob(pattern))
    candidates = [p for p in candidates if p.is_file() and p.suffix.lower() == ".xlsx"]
    if not candidates:
        raise ValueError(f"В папке {directory} не найден файл по маскам: {', '.join(patterns)}")

    def trailing_number(path: Path) -> int:
        match = re.search(r"\((\d+)\)\s*$", path.stem)
        return int(match.group(1)) if match else -1

    # В обычной работе свежий файл имеет самый новый mtime. Номер в скобках — резервный критерий.
    return max(candidates, key=lambda p: (p.stat().st_mtime_ns, trailing_number(p), p.name))


def format_budget(budget_value: Decimal) -> str:
    if budget_value == budget_value.to_integral_value():
        budget = budget_value.quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    else:
        budget = budget_value.normalize()
    return f"{budget}$"


def build_check(
    fb_path: Path,
    mvp_path: Path,
    *,
    warn_mvp_only: bool,
    entity: str = "auto",
) -> str:
    fb_items, resolved_entity = parse_fb(fb_path, entity=entity)
    mvp_items = parse_mvp(mvp_path, entity=resolved_entity)
    entity_label = ENTITY_LABELS[resolved_entity]

    fb_duplicates = duplicate_titles(fb_items)
    mvp_duplicates = duplicate_titles(mvp_items)
    if fb_duplicates or mvp_duplicates:
        blocks = ["⚠️ НЕ СОШЛОСЬ / ПРОВЕРЬ", ""]
        if fb_duplicates:
            blocks += [f"Дублирующиеся названия {entity_label} в ФБ после нормализации пробелов:", *fb_duplicates, ""]
        if mvp_duplicates:
            blocks += [f"Дублирующиеся названия {entity_label} в MVP после нормализации пробелов:", *mvp_duplicates, ""]
        blocks += ["Чек не сформирован: дубли нельзя склеивать наугад."]
        return "\n".join(blocks).rstrip()

    mvp_by_title = {item.normalized_title: item for item in mvp_items}
    fb_by_title = {item.normalized_title: item for item in fb_items}

    fb_only = sorted(item.title for item in fb_items if item.normalized_title not in mvp_by_title)
    mvp_only = sorted(item.title for item in mvp_items if item.normalized_title not in fb_by_title)

    warnings: list[str] = []
    if fb_only or (warn_mvp_only and mvp_only):
        warnings += ["⚠️ НЕ СОШЛОСЬ / ПРОВЕРЬ", ""]
        if fb_only:
            warnings += ["Есть в ФБ, но нет в MVP:", *fb_only, ""]
        if warn_mvp_only and mvp_only:
            warnings += ["Есть в MVP, но нет в ФБ:", *mvp_only, ""]
        warnings += ["Проверь выгрузку.", ""]

    first_seen_geo: dict[str, int] = {}
    grouped: dict[str, list[FbItem]] = defaultdict(list)
    unknown_geo_titles: list[str] = []
    for item in fb_items:
        geo = detect_geo(item.title)
        if geo is None:
            geo = "Неизвестное гео"
            unknown_geo_titles.append(item.title)
        if geo not in first_seen_geo:
            first_seen_geo[geo] = item.first_seen_index
        grouped[geo].append(item)

    if unknown_geo_titles:
        warning_header = ["⚠️ ПРОВЕРЬ", "", "Не удалось определить гео:", *sorted(unknown_geo_titles), "", "Проверь названия.", ""]
        warnings = warning_header + warnings

    priority_index = {geo: idx for idx, geo in enumerate(GEO_PRIORITY)}
    geo_order = sorted(
        grouped,
        key=lambda geo: (0, priority_index[geo]) if geo in priority_index else (1, first_seen_geo[geo]),
    )

    check_lines: list[str] = []
    for geo in geo_order:
        check_lines += [geo, ""]
        total_spend = ZERO
        total_sub = ZERO
        total_chat = ZERO

        # Стандартная строковая сортировка по полному названию из FB, без числовой переупорядочивки.
        for item in sorted(grouped[geo], key=lambda campaign: campaign.title):
            mvp = mvp_by_title.get(item.normalized_title)
            sub = mvp.sub if mvp else ZERO
            chat = mvp.chat if mvp else ZERO
            total_spend += item.spend
            total_sub += sub
            total_chat += chat

            title_line = item.title
            if item.budget is not None:
                title_line += f" [{format_budget(item.budget)}]"
            check_lines += [
                title_line,
                f"{format_money(item.spend)} / {format_metric(item.spend, sub)} / {format_metric(item.spend, chat)}",
                "",
            ]

        check_lines += [
            f"Общ.: {format_money(total_spend)} / {format_metric(total_spend, total_sub)} / {format_metric(total_spend, total_chat)}",
            "",
        ]

    result = warnings + check_lines
    return "\n".join(result).rstrip()


def extract_stats_filter(raw_filter: str) -> str:
    text = normalize_spaces(raw_filter)
    match = re.search(r"(?:общая\s+стата|общая\s+статистика|стата)\s+по\s+(.+)$", text, flags=re.IGNORECASE)
    return normalize_spaces(match.group(1) if match else text)


def split_title_blocks(title: str) -> list[str]:
    # В рабочих названиях смысловые части разделяются именно " - ".
    # Поэтому SPAIN не совпадает с SPAIN-ES, но совпадает с отдельным блоком SPAIN.
    return [normalize_spaces(part).upper() for part in re.split(r"\s+-\s+", title) if normalize_spaces(part)]


def filter_matches(title: str, filter_text: str, *, mode: str) -> bool:
    needle = normalize_spaces(filter_text).upper()
    haystack = normalize_spaces(title).upper()
    if not needle:
        return True
    if mode == "contains":
        return needle in haystack
    if mode == "segment":
        return needle in split_title_blocks(title)
    raise ValueError(f"Некорректный режим фильтра: {mode}")


def build_stats(
    fb_path: Path,
    *,
    raw_filter: str,
    entity: str = "auto",
    filter_mode: str = "segment",
) -> str:
    filter_text = extract_stats_filter(raw_filter)
    fb_items, resolved_entity = parse_fb(
        fb_path,
        entity=entity,
        include_zero_spend=True,
        require_stats=True,
    )
    matched = [item for item in fb_items if filter_matches(item.title, filter_text, mode=filter_mode)]

    total_clicks = sum((item.clicks or ZERO for item in matched), ZERO)
    total_views = sum((item.views or ZERO for item in matched), ZERO)
    total_spend = sum((item.spend for item in matched), ZERO)

    lines = [
        f"Общая стата по {filter_text}",
        f"Уровень: {ENTITY_LABELS[resolved_entity]}",
        f"Найдено строк: {len(matched)}",
        f"Клики: {format_plain_number(total_clicks)}",
        f"Просмотры: {format_plain_number(total_views)}",
        f"Расход: {format_money(total_spend)}",
    ]
    if not matched:
        lines += ["", "⚠️ По фильтру ничего не найдено. По умолчанию фильтр ищет точный блок через ' - '."]
    return "\n".join(lines).rstrip()


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Сформировать текстовый чек из FB и MVP Excel-выгрузок")
    parser.add_argument("--fb", type=Path, help="Путь к FB-выгрузке «Статистика аккаунтов ...xlsx»")
    parser.add_argument("--mvp", type=Path, help="Путь к MVP-выгрузке «data ...xlsx»")
    parser.add_argument("--dir", type=Path, default=Path.cwd(), help="Папка для автоматического выбора свежих файлов")
    parser.add_argument("--output", type=Path, help="Дополнительно сохранить результат в TXT-файл")
    parser.add_argument("--suppress-mvp-only-warning", action="store_true", help="Не показывать предупреждение о строках, которые есть только в MVP")
    parser.add_argument(
        "--entity",
        choices=("auto", "campaign", "ad"),
        default="auto",
        help="Уровень данных: auto, campaign или ad. auto выберет ad, если в FB есть колонка объявления.",
    )
    parser.add_argument(
        "--stats",
        "--stat",
        dest="stats",
        help="Посчитать общие клики, просмотры и расход по фильтру. Например: --stats SPAIN",
    )
    parser.add_argument(
        "--filter-mode",
        choices=("segment", "contains"),
        default="segment",
        help="segment = точный блок через ' - '; contains = обычное вхождение текста.",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        fb_path = args.fb or auto_pick(args.dir, ("Статистика аккаунтов*.xlsx", "Статистика аккаунта*.xlsx"))
        if not fb_path.exists():
            raise ValueError(f"FB-файл не найден: {fb_path}")

        if args.stats:
            result = build_stats(
                fb_path,
                raw_filter=args.stats,
                entity=args.entity,
                filter_mode=args.filter_mode,
            )
        else:
            mvp_path = args.mvp or auto_pick(args.dir, ("data*.xlsx",))
            if not mvp_path.exists():
                raise ValueError(f"MVP-файл не найден: {mvp_path}")
            result = build_check(
                fb_path,
                mvp_path,
                warn_mvp_only=not args.suppress_mvp_only_warning,
                entity=args.entity,
            )
    except Exception as exc:
        print(f"⚠️ ПРОВЕРЬ\n\nФайл не читается или структура выгрузки некорректна:\n{exc}", file=sys.stderr)
        return 1

    print(result)
    if args.output:
        args.output.write_text(result + "\n", encoding="utf-8")
    return 0


# =========================
# Streamlit app layer
# =========================

def build_check_from_items(
    fb_items: list[FbItem],
    mvp_items: list[MvpItem],
    *,
    warn_mvp_only: bool,
    resolved_entity: str,
) -> str:
    """Та же сборка чека, что и build_check(), но без повторного чтения Excel."""
    entity_label = ENTITY_LABELS[resolved_entity]

    fb_duplicates = duplicate_titles(fb_items)
    mvp_duplicates = duplicate_titles(mvp_items)
    if fb_duplicates or mvp_duplicates:
        blocks = ["⚠️ НЕ СОШЛОСЬ / ПРОВЕРЬ", ""]
        if fb_duplicates:
            blocks += [f"Дублирующиеся названия {entity_label} в ФБ после нормализации пробелов:", *fb_duplicates, ""]
        if mvp_duplicates:
            blocks += [f"Дублирующиеся названия {entity_label} в MVP после нормализации пробелов:", *mvp_duplicates, ""]
        blocks += ["Чек не сформирован: дубли нельзя склеивать наугад."]
        return "\n".join(blocks).rstrip()

    mvp_by_title = {item.normalized_title: item for item in mvp_items}
    fb_by_title = {item.normalized_title: item for item in fb_items}

    fb_only = sorted(item.title for item in fb_items if item.normalized_title not in mvp_by_title)
    mvp_only = sorted(item.title for item in mvp_items if item.normalized_title not in fb_by_title)

    warnings: list[str] = []
    if fb_only or (warn_mvp_only and mvp_only):
        warnings += ["⚠️ НЕ СОШЛОСЬ / ПРОВЕРЬ", ""]
        if fb_only:
            warnings += ["Есть в ФБ, но нет в MVP:", *fb_only, ""]
        if warn_mvp_only and mvp_only:
            warnings += ["Есть в MVP, но нет в ФБ:", *mvp_only, ""]
        warnings += ["Проверь выгрузку.", ""]

    first_seen_geo: dict[str, int] = {}
    grouped: dict[str, list[FbItem]] = defaultdict(list)
    unknown_geo_titles: list[str] = []
    for item in fb_items:
        geo = detect_geo(item.title)
        if geo is None:
            geo = "Неизвестное гео"
            unknown_geo_titles.append(item.title)
        if geo not in first_seen_geo:
            first_seen_geo[geo] = item.first_seen_index
        grouped[geo].append(item)

    if unknown_geo_titles:
        warning_header = ["⚠️ ПРОВЕРЬ", "", "Не удалось определить гео:", *sorted(unknown_geo_titles), "", "Проверь названия.", ""]
        warnings = warning_header + warnings

    priority_index = {geo: idx for idx, geo in enumerate(GEO_PRIORITY)}
    geo_order = sorted(
        grouped,
        key=lambda geo: (0, priority_index[geo]) if geo in priority_index else (1, first_seen_geo[geo]),
    )

    check_lines: list[str] = []
    for geo in geo_order:
        check_lines += [geo, ""]
        total_spend = ZERO
        total_sub = ZERO
        total_chat = ZERO

        for item in sorted(grouped[geo], key=lambda campaign: campaign.title):
            mvp = mvp_by_title.get(item.normalized_title)
            sub = mvp.sub if mvp else ZERO
            chat = mvp.chat if mvp else ZERO
            total_spend += item.spend
            total_sub += sub
            total_chat += chat

            title_line = item.title
            if item.budget is not None:
                title_line += f" [{format_budget(item.budget)}]"
            check_lines += [
                title_line,
                f"{format_money(item.spend)} / {format_metric(item.spend, sub)} / {format_metric(item.spend, chat)}",
                "",
            ]

        check_lines += [
            f"Общ.: {format_money(total_spend)} / {format_metric(total_spend, total_sub)} / {format_metric(total_spend, total_chat)}",
            "",
        ]

    result = warnings + check_lines
    return "\n".join(result).rstrip()


def extract_title_meta(title: str) -> dict[str, str]:
    parts = [normalize_spaces(part) for part in re.split(r"\s+-\s+", title) if normalize_spaces(part)]

    date_value = ""
    for part in parts:
        match = re.search(r"\b\d{1,2}[./-]\d{1,2}(?:[./-]\d{2,4})?\b", part)
        if match:
            date_value = match.group(0)
            break

    edit_value = ""
    for part in parts:
        match = re.search(r"\bEDIT\s*\d+\b", part, flags=re.IGNORECASE)
        if match:
            edit_value = re.sub(r"\s+", "", match.group(0).upper())
            break

    cabinet_value = ""
    for part in parts:
        match = re.search(r"\b\d*T\d+A\b", part, flags=re.IGNORECASE)
        if match:
            cabinet_value = match.group(0).upper()
            break

    geo_value = detect_geo(title) or "Неизвестное гео"
    return {
        "date": date_value,
        "geo": geo_value,
        "edit": edit_value,
        "cabinet": cabinet_value,
        "t2a": cabinet_value,
    }


def decimal_to_float(value: Decimal | None) -> float | None:
    if value is None:
        return None
    return float(value)


def decimal_to_int_or_float(value: Decimal | None) -> int | float | None:
    if value is None:
        return None
    if value == value.to_integral_value():
        return int(value)
    return float(value)


def build_app_context(
    fb_path: Path,
    mvp_path: Path | None,
    *,
    entity: str,
    warn_mvp_only: bool,
) -> dict[str, object]:
    fb_items, resolved_entity = parse_fb(fb_path, entity=entity, include_zero_spend=False)
    mvp_items = parse_mvp(mvp_path, entity=resolved_entity) if mvp_path else []

    check_text = ""
    if mvp_path:
        check_text = build_check_from_items(
            fb_items,
            mvp_items,
            warn_mvp_only=warn_mvp_only,
            resolved_entity=resolved_entity,
        )

    mvp_by_title = {item.normalized_title: item for item in mvp_items}
    fb_by_title = {item.normalized_title: item for item in fb_items}
    rows: list[dict[str, object]] = []

    for item in fb_items:
        mvp = mvp_by_title.get(item.normalized_title)
        sub = mvp.sub if mvp else ZERO
        chat = mvp.chat if mvp else ZERO
        meta = extract_title_meta(item.title)
        pdp_price = None if sub == 0 else item.spend / sub
        dialog_price = None if chat == 0 else item.spend / chat
        rows.append(
            {
                "Статус": "✅ OK" if mvp else "⚠️ Есть в ФБ, нет в MVP",
                "Название": item.title,
                "Гео": meta["geo"],
                "Дата": meta["date"],
                "EDIT": meta["edit"],
                "Кабинет/T2A": meta["cabinet"],
                "Бюджет": format_budget(item.budget) if item.budget is not None else "",
                "Расход": decimal_to_float(item.spend),
                "ПДП": decimal_to_int_or_float(sub),
                "Диа": decimal_to_int_or_float(chat),
                "Цена ПДП": decimal_to_float(pdp_price),
                "Цена диа": decimal_to_float(dialog_price),
                "Клики": decimal_to_int_or_float(item.clicks),
                "Просмотры": decimal_to_int_or_float(item.views),
                "FB строка": item.row_number,
                "MVP строка": mvp.row_number if mvp else None,
                "В чеке": True,
            }
        )

    if mvp_path:
        for item in mvp_items:
            if item.normalized_title in fb_by_title:
                continue
            meta = extract_title_meta(item.title)
            rows.append(
                {
                    "Статус": "⚠️ Есть в MVP, нет в ФБ",
                    "Название": item.title,
                    "Гео": meta["geo"],
                    "Дата": meta["date"],
                    "EDIT": meta["edit"],
                    "Кабинет/T2A": meta["cabinet"],
                    "Бюджет": "",
                    "Расход": 0.0,
                    "ПДП": decimal_to_int_or_float(item.sub),
                    "Диа": decimal_to_int_or_float(item.chat),
                    "Цена ПДП": None,
                    "Цена диа": None,
                    "Клики": None,
                    "Просмотры": None,
                    "FB строка": None,
                    "MVP строка": item.row_number,
                    "В чеке": False,
                }
            )

    fb_duplicates = duplicate_titles(fb_items)
    mvp_duplicates = duplicate_titles(mvp_items) if mvp_path else []
    return {
        "rows": rows,
        "check_text": check_text,
        "resolved_entity": resolved_entity,
        "fb_count": len(fb_items),
        "mvp_count": len(mvp_items),
        "fb_duplicates": fb_duplicates,
        "mvp_duplicates": mvp_duplicates,
    }


def app_filter_dataframe(df, *, search_text: str, geos: list[str], dates: list[str], edits: list[str], cabinets: list[str]):
    filtered = df.copy()
    if search_text:
        needle = normalize_spaces(search_text).upper()
        filtered = filtered[filtered["Название"].fillna("").str.upper().str.contains(re.escape(needle), regex=True)]
    if geos:
        filtered = filtered[filtered["Гео"].isin(geos)]
    if dates:
        filtered = filtered[filtered["Дата"].isin(dates)]
    if edits:
        filtered = filtered[filtered["EDIT"].isin(edits)]
    if cabinets:
        filtered = filtered[filtered["Кабинет/T2A"].isin(cabinets)]
    return filtered


def app_highlight_mismatches(row):
    status = str(row.get("Статус", ""))
    if "ФБ, нет в MVP" in status:
        return ["background-color: #fff3cd"] * len(row)
    if "MVP, нет в ФБ" in status:
        return ["background-color: #f8d7da"] * len(row)
    return [""] * len(row)


def app_copy_button(text: str, *, label: str = "Скопировать чек", key: str = "copy-check") -> None:
    import html
    import json
    import streamlit.components.v1 as components

    safe_key = re.sub(r"[^a-zA-Z0-9_-]", "-", key)
    label_html = html.escape(label)
    text_json = json.dumps(text)
    components.html(
        f"""
        <button id="{safe_key}" style="
            width: 100%;
            padding: 0.65rem 0.9rem;
            border: 1px solid #d0d7de;
            border-radius: 0.5rem;
            background: #ffffff;
            cursor: pointer;
            font-size: 0.95rem;
        ">📋 {label_html}</button>
        <script>
        const btn = document.getElementById({json.dumps(safe_key)});
        const sourceText = {text_json};
        btn.addEventListener('click', async () => {{
            try {{
                await navigator.clipboard.writeText(sourceText);
                btn.innerText = '✅ Скопировано';
            }} catch (err) {{
                const textarea = document.createElement('textarea');
                textarea.value = sourceText;
                textarea.style.position = 'fixed';
                textarea.style.opacity = '0';
                document.body.appendChild(textarea);
                textarea.focus();
                textarea.select();
                document.execCommand('copy');
                document.body.removeChild(textarea);
                btn.innerText = '✅ Скопировано';
            }}
            setTimeout(() => {{ btn.innerText = '📋 {label_html}'; }}, 1400);
        }});
        </script>
        """,
        height=56,
    )


def app_money_metric(value: float | int | None) -> str:
    if value is None:
        return "—"
    return f"${float(value):.2f}"


def app_number_metric(value: float | int | None) -> str:
    if value is None:
        return "—"
    value_float = float(value)
    if value_float.is_integer():
        return str(int(value_float))
    return f"{value_float:.2f}"


def write_upload_to_temp(uploaded_file, directory: Path, fallback_name: str) -> Path:
    safe_name = re.sub(r"[^\w.() -]+", "_", uploaded_file.name or fallback_name)
    if not safe_name.lower().endswith(".xlsx"):
        safe_name += ".xlsx"
    path = directory / safe_name
    path.write_bytes(uploaded_file.getvalue())
    return path


def run_streamlit_app() -> None:
    import tempfile
    import pandas as pd
    import streamlit as st

    st.set_page_config(page_title="Forex Check", page_icon="📊", layout="wide")
    st.title("📊 Forex Check")
    st.caption("Загрузи FB и MVP Excel — приложение соберёт чек, покажет расхождения и посчитает суммы по фильтрам.")

    with st.sidebar:
        st.header("Файлы")
        fb_upload = st.file_uploader("FB выгрузка / Статистика аккаунтов", type=["xlsx"], key="fb_upload")
        mvp_upload = st.file_uploader("MVP выгрузка / data", type=["xlsx"], key="mvp_upload")

        st.header("Настройки")
        entity = st.radio(
            "Уровень расчёта",
            options=("auto", "campaign", "ad"),
            index=0,
            format_func=lambda value: {
                "auto": "Auto: объявления, если есть колонка объявлений",
                "campaign": "Кампании",
                "ad": "Объявления",
            }[value],
        )
        warn_mvp_only = st.checkbox("Показывать строки, которые есть только в MVP", value=True)

    if not fb_upload:
        st.info("Загрузи FB-файл, чтобы начать. Для полного чека нужен ещё MVP-файл.")
        return

    try:
        with tempfile.TemporaryDirectory() as tmp_dir_raw:
            tmp_dir = Path(tmp_dir_raw)
            fb_path = write_upload_to_temp(fb_upload, tmp_dir, "fb.xlsx")
            mvp_path = write_upload_to_temp(mvp_upload, tmp_dir, "mvp.xlsx") if mvp_upload else None
            context = build_app_context(
                fb_path,
                mvp_path,
                entity=entity,
                warn_mvp_only=warn_mvp_only,
            )
    except Exception as exc:
        st.error(f"Файл не читается или структура выгрузки некорректна: {exc}")
        return

    rows = context["rows"]
    df = pd.DataFrame(rows)
    resolved_entity = str(context["resolved_entity"])
    check_text = str(context["check_text"] or "")

    st.success(f"Файлы прочитаны. Уровень: {'объявления' if resolved_entity == 'ad' else 'кампании'}.")

    if context["fb_duplicates"] or context["mvp_duplicates"]:
        with st.expander("⚠️ Дубли после нормализации пробелов", expanded=True):
            if context["fb_duplicates"]:
                st.write("**ФБ:**")
                st.write(context["fb_duplicates"])
            if context["mvp_duplicates"]:
                st.write("**MVP:**")
                st.write(context["mvp_duplicates"])
            st.warning("При дублях чек не склеивается наугад. Лучше поправить выгрузки/названия.")

    tab_check, tab_filter, tab_search, tab_mismatch = st.tabs(
        ["Готовый чек", "Сумма по фильтру", "Быстрый поиск", "Несовпадения FB/MVP"]
    )

    with tab_check:
        if not mvp_upload:
            st.warning("Для готового чека загрузи MVP-файл. Блок суммы по фильтру может работать только по FB, если в FB есть клики/просмотры/расход.")
        else:
            col_copy, col_download = st.columns([1, 1])
            with col_copy:
                app_copy_button(check_text, label="Скопировать чек", key="copy-check-main")
            with col_download:
                st.download_button(
                    "⬇️ Скачать TXT",
                    data=check_text + "\n",
                    file_name="check.txt",
                    mime="text/plain",
                    use_container_width=True,
                )
            st.text_area("Чек", value=check_text, height=520)

    with tab_filter:
        st.subheader("Посчитать сумму по фильтру")
        st.caption("Фильтры можно комбинировать: название, гео, дата и кабинет/T2A. EDIT можно искать через поле «Название содержит».")

        col1, col2, col3 = st.columns(3)
        with col1:
            search_text = st.text_input("Название содержит", placeholder="Например: SPAIN, EDIT1, 6T2A")
            geo_values = sorted(x for x in df["Гео"].dropna().unique().tolist() if x)
            geos = st.multiselect("Гео", geo_values)
        with col2:
            date_values = sorted(x for x in df["Дата"].dropna().unique().tolist() if x)
            dates = st.multiselect("Дата", date_values)
            edits = []
        with col3:
            cabinet_values = sorted(x for x in df["Кабинет/T2A"].dropna().unique().tolist() if x)
            cabinets = st.multiselect("Кабинет / T2A", cabinet_values)
            only_check_rows = st.checkbox("Считать только строки из чека / FB со спендом", value=True)

        filtered = app_filter_dataframe(
            df,
            search_text=search_text,
            geos=geos,
            dates=dates,
            edits=edits,
            cabinets=cabinets,
        )
        if only_check_rows and "В чеке" in filtered.columns:
            filtered = filtered[filtered["В чеке"] == True]

        spend_sum = float(filtered["Расход"].fillna(0).sum()) if "Расход" in filtered else 0.0
        clicks_sum = filtered["Клики"].dropna().sum() if "Клики" in filtered and filtered["Клики"].notna().any() else None
        views_sum = filtered["Просмотры"].dropna().sum() if "Просмотры" in filtered and filtered["Просмотры"].notna().any() else None
        sub_sum = filtered["ПДП"].fillna(0).sum() if "ПДП" in filtered else 0
        chat_sum = filtered["Диа"].fillna(0).sum() if "Диа" in filtered else 0

        metric_cols = st.columns(6)
        metric_cols[0].metric("Найдено строк", len(filtered))
        metric_cols[1].metric("Расход", app_money_metric(spend_sum))
        metric_cols[2].metric("Клики", app_number_metric(clicks_sum))
        metric_cols[3].metric("Просмотры", app_number_metric(views_sum))
        metric_cols[4].metric("ПДП", app_number_metric(sub_sum))
        metric_cols[5].metric("Диа", app_number_metric(chat_sum))

        cpa_cols = st.columns(2)
        cpa_cols[0].metric("Общая цена ПДП", "0" if not sub_sum else app_money_metric(spend_sum / float(sub_sum)))
        cpa_cols[1].metric("Общая цена диа", "0" if not chat_sum else app_money_metric(spend_sum / float(chat_sum)))

        show_cols = [
            "Статус", "Название", "Гео", "Дата", "Кабинет/T2A", "Бюджет",
            "Расход", "ПДП", "Диа", "Цена ПДП", "Цена диа", "Клики", "Просмотры",
        ]
        show_cols = [col for col in show_cols if col in filtered.columns]
        st.dataframe(
            filtered[show_cols].style.apply(app_highlight_mismatches, axis=1),
            use_container_width=True,
            hide_index=True,
        )

        export_text = "\n".join(filtered["Название"].astype(str).tolist())
        if export_text:
            app_copy_button(export_text, label="Скопировать найденные названия", key="copy-filtered-names")

    with tab_search:
        st.subheader("Быстрый поиск кампаний / объявлений")
        quick_query = st.text_input("Поиск", placeholder="Введи часть названия, гео, EDIT или T2A", key="quick_search")
        quick_filtered = app_filter_dataframe(
            df,
            search_text=quick_query,
            geos=[],
            dates=[],
            edits=[],
            cabinets=[],
        ) if quick_query else df.head(100)
        st.caption(f"Показано строк: {len(quick_filtered)}")
        quick_cols = ["Статус", "Название", "Гео", "Дата", "EDIT", "Кабинет/T2A", "Расход", "ПДП", "Диа", "Клики", "Просмотры"]
        quick_cols = [col for col in quick_cols if col in quick_filtered.columns]
        st.dataframe(
            quick_filtered[quick_cols].style.apply(app_highlight_mismatches, axis=1),
            use_container_width=True,
            hide_index=True,
        )

    with tab_mismatch:
        st.subheader("Подсветка несовпадений FB/MVP")
        mismatches = df[df["Статус"] != "✅ OK"] if "Статус" in df else df.iloc[0:0]
        if mismatches.empty:
            st.success("Несовпадений не найдено.")
        else:
            st.warning(f"Найдено несовпадений: {len(mismatches)}")
            mismatch_cols = ["Статус", "Название", "Гео", "Дата", "EDIT", "Кабинет/T2A", "Расход", "ПДП", "Диа", "FB строка", "MVP строка"]
            mismatch_cols = [col for col in mismatch_cols if col in mismatches.columns]
            st.dataframe(
                mismatches[mismatch_cols].style.apply(app_highlight_mismatches, axis=1),
                use_container_width=True,
                hide_index=True,
            )


def _running_inside_streamlit() -> bool:
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx
        return get_script_run_ctx() is not None
    except Exception:
        return False


if __name__ == "__main__":
    if _running_inside_streamlit():
        run_streamlit_app()
    else:
        # Старый режим тоже оставлен: файл можно запускать через python как CLI-скрипт.
        raise SystemExit(main())
